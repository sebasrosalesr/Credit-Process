# streamlit_summary_totals.py
import re
import pandas as pd
import streamlit as st
from openpyxl import load_workbook

st.set_page_config(page_title="Summary Totals Extractor", layout="wide")
st.title("📊 Summary Totals Extractor")
st.write(
    "Upload an Excel workbook with multiple sheets. "
    "The app will detect lines such as **'Summary total for customer …'** "
    "and extract the actual totals beside or below them."
)

# -------------------------
# File upload
# -------------------------
uploaded_file = st.file_uploader("📤 Upload Excel file", type=["xlsx", "xlsm", "xls"])

if uploaded_file:
    st.info("Processing all sheets... this may take a few seconds.")
    wb = load_workbook(uploaded_file, data_only=True)
    records = []

    def find_rightward_number(row_cells, start_col_idx):
        """Scan to the right for a number or currency value."""
        for j in range(start_col_idx + 1, min(start_col_idx + 8, len(row_cells))):
            v = row_cells[j].value
            if isinstance(v, (int, float)) and v != 0:
                return v, row_cells[j].coordinate
            if isinstance(v, str):
                m = re.search(r"\$?\s?(\d[\d,]*\.?\d*)", v)
                if m:
                    return float(m.group(1).replace(",", "")), row_cells[j].coordinate
        return None, None

    for ws in wb.worksheets:
        try:
            for r in ws.iter_rows(values_only=False):
                for c_idx, cell in enumerate(r):
                    val = cell.value
                    if isinstance(val, str) and "summary total for customer" in val.lower():
                        num, coord = find_rightward_number(r, c_idx)
                        if num is None:
                            # if not found to the right, check next 3 rows same column
                            for k in range(1, 4):
                                rr = cell.row + k
                                if rr <= ws.max_row:
                                    v2 = ws.cell(row=rr, column=cell.column).value
                                    if isinstance(v2, (int, float)) and v2 != 0:
                                        num, coord = v2, ws.cell(row=rr, column=cell.column).coordinate
                                        break
                                    if isinstance(v2, str):
                                        m = re.search(r"\$?\s?(\d[\d,]*\.?\d*)", v2)
                                        if m:
                                            num = float(m.group(1).replace(",", ""))
                                            coord = ws.cell(row=rr, column=cell.column).coordinate
                                            break
                        records.append({
                            "Sheet": ws.title,
                            "Cell (Text)": cell.coordinate,
                            "Summary Text": val,
                            "Total": num,
                            "Cell (Total)": coord
                        })
                        break  # move to next row once found
        except Exception as e:
            records.append({"Sheet": ws.title, "Error": str(e)})

    if records:
        df = pd.DataFrame(records)
        st.success(f"✅ Found {len(df)} summary totals across {len(wb.worksheets)} sheets.")
        st.dataframe(df, use_container_width=True)

        # Create cleaned output for XLOOKUP
        df_clean = df.copy()
        df_clean["Customer Code"] = df_clean["Summary Text"].str.extract(r"customer\s+([A-Za-z0-9\/\-]+)", expand=False)
        df_clean["Total Clean"] = pd.to_numeric(df_clean["Total"], errors="coerce")

        csv = df_clean.to_csv(index=False).encode("utf-8")
        st.download_button(
            "⬇️ Download CSV with Totals",
            data=csv,
            file_name="summary_totals_extracted.csv",
            mime="text/csv",
        )
    else:
        st.warning("No 'Summary total for customer' lines were found in this workbook.")
else:
    st.info("Please upload an Excel workbook to begin.")
